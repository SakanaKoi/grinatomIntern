import pandas as pd
import openpyxl
import os
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter


def initialize_excel_file(file_path):
    # Удаление Excel файла если есть
    if os.path.exists(file_path):
        os.remove(file_path)

    # Создание нового Excel файла
    headers = ['Дата USD', 'Курс USD', 'Время USD', 'Дата JPY', 'Курс JPY', 'Время JPY', 'Результат']
    df = pd.DataFrame(columns=headers)
    df.to_excel(file_path, index=False)


# Запись данных о курсе USD в Excel
def write_usd_data_to_excel(data, file_path):
    try:
        df = pd.DataFrame(data, columns=['Дата USD', 'Курс USD', 'Время USD'])
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=1)
    except Exception as e:
        raise Exception("Ошибка при записи данных о USD в excel") from e


# Запись данных о курсе JPY в Excel
def write_jpy_data_to_excel(data, file_path):
    try:
        df = pd.DataFrame(data, columns=['Дата JPY', 'Курс JPY', 'Время JPY'])
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=1, startcol=3)
    except Exception as e:
        raise Exception("Ошибка при записи данных о JPY в excel") from e


# Вычисление частного курса USD к JPY и запись в столбец 'Результат'
def calculate_usd_to_jpy_ratio(file_path):
    try:
        df = pd.read_excel(file_path)

        df['Результат'] = df.apply(
            lambda row: row['Курс USD'] / row['Курс JPY'] if pd.notnull(row['Курс USD']) and pd.notnull(
                row['Курс JPY']) else None, axis=1)

        df.to_excel(file_path, index=False)
    except Exception as e:
        raise Exception("Ошибка при расчете частного курса USD к JPY") from e


# Форматирование таблицы Excel
def format_excel_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Создание стилей
        financial_style = NamedStyle(name="financial", number_format="#,##0.00 ₽")
        date_style = NamedStyle(name="date", number_format="DD.MM.YYYY")
        time_style = NamedStyle(name="time", number_format="HH:MM:SS")

        # Применение стилей
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            row[1].style = financial_style  # Курс USD
            row[4].style = financial_style  # Курс JPY
            row[0].style = date_style  # Дата USD
            row[3].style = date_style  # Дата JPY
            row[2].style = time_style  # Время USD
            row[5].style = time_style  # Время JPY

        # Автоподбор ширины ячеек
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
    except Exception as e:
        raise Exception("Ошибка при форматировании Excel файла") from e
